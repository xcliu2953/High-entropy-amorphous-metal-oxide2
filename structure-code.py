import openpyxl

# 打开Excel文件
workbook = openpyxl.load_workbook('matched_letters.xlsx')
sheet = workbook.active

# 定义字母与替换值的映射字典
letter_to_value = {
    'A': 1,
    'B': 2,
    'C': 3,
    'D': 4,
    'E': 5,
    'F': 6,
    'G': 7,
    'H': 8,
    'I': 9,
    'J': 10,
    'K': 11,
    'L': 12,
    'M': 13,
    'N': 14,
    'O': 15,
    'P': 16,
}

# 遍历Excel中的每一行并进行替换
for row in sheet.iter_rows(min_row=2, values_only=True):
    for col_index in range(2, 18):  # 从第二列到第17列
        cell_value = row[col_index - 2]  # 由于列表索引从0开始，所以需要减2
        if cell_value in letter_to_value:
            updated_value = letter_to_value[cell_value]
            sheet.cell(row=row[0].row, column=col_index, value=updated_value)

# 保存修改后的Excel文件
workbook.save('matched_letters_replaced.xlsx')
workbook.close()
